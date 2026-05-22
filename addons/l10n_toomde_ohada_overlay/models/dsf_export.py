"""DSF + Form 1004 XLSX exports for OHADA tenants.

Two export families:

* DSF = Déclaration Statistique et Fiscale. Per-country annual
  declaration that combines the financial statements with tax-base
  schedules used by the tax authority for cross-checks. Each member
  state publishes its own template; the structure of the data is the
  same SYSCOHADA roll-up but cell positions and column labels differ.

* Form 1004 = OHADA-wide standardized financial-statement filing form
  attached to the Acte Uniforme. Same content for every member state.

Class hierarchy:

    BaseOhadaAnnualExport         (shared XLSX builder)
      ├── DsfSenegalExport        — COMP-9 (SN/DGID)
      ├── DsfCoteIvoireExport     — COMP-9 (CI/DGI)
      ├── DsfBurkinaFasoExport    — COMP-9 (BF/DGI)
      ├── DsfMaliExport           — COMP-9 (ML/DGI)
      ├── DsfCamerounExport       — COMP-9 (CM/DGI)
      └── Form1004Export          — COMP-10 (OHADA-wide)

Per-country subclasses override `_rows()` to inject country-specific
labels and any extra schedules (CI's BIC withholding declaration,
Cameroon's CFA-zone reconciliation, etc.). The class digit roll-up
from `_class_totals()` is shared.

The XLSX layout is still "Toomde first-pass" — column labels and cell
positions will be aligned with each DGI's published template during
the COMP-11 attestation engagement. Auditors typically dictate exact
cell positions; pre-aligning blind produces churn.
"""

from __future__ import annotations

import io
import logging
from datetime import date

from odoo import api, models, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


# Country → DSF model. Used by the wizard to pick the right subclass.
# One entry per OHADA member state with a published annual DSF
# (Déclaration Statistique et Fiscale) filing format. Note that the
# SYSCOHADA chart is uniform across OHADA, so the *body* of every DSF
# is the same Class 1-8 roll-up; what differs is the title, the
# authority name (DGI / DGID / OTR / etc.), and country-specific tax
# lines (BIC withholding in CI, IUTS in BF, …).
DSF_BY_COUNTRY = {
    # West Africa — UEMOA
    "SN": "toomde.ohada.dsf.sn",  # DGID Sénégal
    "CI": "toomde.ohada.dsf.ci",  # DGI Côte d'Ivoire
    "BF": "toomde.ohada.dsf.bf",  # DGI Burkina Faso
    "ML": "toomde.ohada.dsf.ml",  # DGI Mali
    "NE": "toomde.ohada.dsf.ne",  # DGI Niger
    "TG": "toomde.ohada.dsf.tg",  # OTR Togo
    "BJ": "toomde.ohada.dsf.bj",  # DGI Bénin
    "GW": "toomde.ohada.dsf.gw",  # DGCI Guinée-Bissau
    # Central Africa — CEMAC + RDC
    "CM": "toomde.ohada.dsf.cm",  # DGI Cameroun
    "GA": "toomde.ohada.dsf.ga",  # DGI Gabon
    "CD": "toomde.ohada.dsf.cd",  # DGI RDC
}


class BaseOhadaAnnualExport(models.AbstractModel):
    _name = "toomde.ohada.annual.export.base"
    _description = "Base — annual SYSCOHADA export builder"

    @api.model
    def _class_totals(self, company, year):
        """Sum of `account.move.line.balance` per Class 1..8 prefix.

        Filters: only posted moves, only in the fiscal year `year`
        (calendar year — fiscal-year-end customization is a follow-up).
        Returns a dict keyed by class digit -> float.
        """
        start = date(year, 1, 1)
        end = date(year, 12, 31)
        line_model = self.env["account.move.line"]
        totals = {str(i): 0.0 for i in range(1, 9)}
        for class_digit in totals:
            balance = sum(
                line_model.search([
                    ("company_id", "=", company.id),
                    ("parent_state", "=", "posted"),
                    ("date", ">=", start),
                    ("date", "<=", end),
                    ("account_id.code", "=like", f"{class_digit}%"),
                ]).mapped("balance"))
            totals[class_digit] = balance
        return totals

    @api.model
    def _prefix_total(self, company, year, prefix):
        """Sum of balance for accounts whose code starts with `prefix`
        in the given fiscal year. Used by per-country DSF subclasses
        that need finer-grained subtotals than the class-level digest."""
        start = date(year, 1, 1)
        end = date(year, 12, 31)
        return sum(
            self.env["account.move.line"].search([
                ("company_id", "=", company.id),
                ("parent_state", "=", "posted"),
                ("date", ">=", start),
                ("date", "<=", end),
                ("account_id.code", "=like", f"{prefix}%"),
            ]).mapped("balance"))

    @api.model
    def _rows(self, company, year):
        """Country-specific rows. Subclass overrides.

        Returns a list of tuples: `(label, value)`."""
        raise NotImplementedError

    @api.model
    def xlsx(self, company, year) -> bytes:
        """Render the export as an XLSX workbook (bytes)."""
        try:
            import openpyxl
        except ImportError:
            raise UserError(_("openpyxl is required for the SYSCOHADA exports."))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self._sheet_name()

        ws["A1"] = self._title()
        ws["A2"] = f"{company.name} — {year}"
        ws["A3"] = f"Pays: {company.country_id.name or ''}"
        if company.vat:
            ws["A4"] = f"NIF: {company.vat}"

        # Roll-up table starts at row 6.
        ws["A6"] = "Libellé"
        ws["B6"] = f"Exercice {year}"

        for i, (label, value) in enumerate(self._rows(company, year), start=7):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _title(self):
        raise NotImplementedError

    def _sheet_name(self):
        raise NotImplementedError


# ---------------------------------------------------------------------------
# Shared rows for OHADA-zone DSFs (SYSCOHADA chart is the same across
# countries; the per-country subclass overrides the title and any
# country-specific extra schedules — see CI's BIC withholding below).
# ---------------------------------------------------------------------------


def _ohada_common_rows(model, company, year):
    """Class 1-8 roll-up — the universal SYSCOHADA backbone every DSF
    in the OHADA zone needs."""
    totals = model._class_totals(company, year)
    return [
        ("Classe 1 — Capitaux et ressources assimilées", -totals["1"]),
        ("Classe 2 — Immobilisations", totals["2"]),
        ("Classe 3 — Stocks et en-cours", totals["3"]),
        ("Classe 4 — Tiers", totals["4"]),
        ("Classe 5 — Trésorerie", totals["5"]),
        ("Classe 6 — Charges", totals["6"]),
        ("Classe 7 — Produits", -totals["7"]),
        ("Classe 8 — Résultat HAO", -totals["8"]),
    ]


# ---------------------------------------------------------------------------
# Per-country DSF — Sénégal (DGID)
# ---------------------------------------------------------------------------

class DsfSenegalExport(models.AbstractModel):
    _name = "toomde.ohada.dsf.sn"
    _inherit = "toomde.ohada.annual.export.base"
    _description = "DSF Sénégal annual export (COMP-9 — DGID)"

    def _title(self):
        return "Déclaration Statistique et Fiscale — Sénégal (DGID)"

    def _sheet_name(self):
        return "DSF Sénégal"

    @api.model
    def _rows(self, company, year):
        rows = _ohada_common_rows(self, company, year)
        # DGID requests specific SYSCOHADA sub-totals — TVA, CGU, IS.
        rows.append(("", ""))
        rows.append(("TVA collectée (compte 4431)", -self._prefix_total(company, year, "4431")))
        rows.append(("TVA déductible (compte 4452)", self._prefix_total(company, year, "4452")))
        rows.append(("Impôt sur les sociétés (compte 891)", self._prefix_total(company, year, "891")))
        return rows


# ---------------------------------------------------------------------------
# Per-country DSF — Côte d'Ivoire (DGI)
# ---------------------------------------------------------------------------

class DsfCoteIvoireExport(models.AbstractModel):
    _name = "toomde.ohada.dsf.ci"
    _inherit = "toomde.ohada.annual.export.base"
    _description = "DSF Côte d'Ivoire annual export (COMP-9 — DGI CI)"

    def _title(self):
        return "Déclaration Statistique et Fiscale — Côte d'Ivoire (DGI)"

    def _sheet_name(self):
        return "DSF Côte d'Ivoire"

    @api.model
    def _rows(self, company, year):
        rows = _ohada_common_rows(self, company, year)
        # CI DGI declaration includes BIC withholding (7.5% on services
        # from non-domiciliated suppliers) — separated from the generic
        # withholding line so DGI can cross-check.
        rows.append(("", ""))
        rows.append(("TVA collectée (compte 4431)", -self._prefix_total(company, year, "4431")))
        rows.append(("TVA déductible (compte 4452)", self._prefix_total(company, year, "4452")))
        rows.append(("Retenues BIC sur prestations (compte 4424)", -self._prefix_total(company, year, "4424")))
        rows.append(("Impôt BIC (compte 891)", self._prefix_total(company, year, "891")))
        return rows


# ---------------------------------------------------------------------------
# Per-country DSF — Burkina Faso (DGI)
# ---------------------------------------------------------------------------

class DsfBurkinaFasoExport(models.AbstractModel):
    _name = "toomde.ohada.dsf.bf"
    _inherit = "toomde.ohada.annual.export.base"
    _description = "DSF Burkina Faso annual export (COMP-9 — DGI BF)"

    def _title(self):
        return "Déclaration Statistique et Fiscale — Burkina Faso (DGI)"

    def _sheet_name(self):
        return "DSF Burkina Faso"

    @api.model
    def _rows(self, company, year):
        rows = _ohada_common_rows(self, company, year)
        # BF DGI declaration includes IUTS (Impôt Unique sur les
        # Traitements et Salaires, 5%) reconciliation alongside the
        # standard TVA section.
        rows.append(("", ""))
        rows.append(("TVA collectée (compte 4431)", -self._prefix_total(company, year, "4431")))
        rows.append(("TVA déductible (compte 4452)", self._prefix_total(company, year, "4452")))
        rows.append(("Retenues IUTS sur salaires (compte 4423)", -self._prefix_total(company, year, "4423")))
        rows.append(("Impôt sur les sociétés (compte 891)", self._prefix_total(company, year, "891")))
        return rows


# ---------------------------------------------------------------------------
# Per-country DSF — Mali (DGI)
# ---------------------------------------------------------------------------

class DsfMaliExport(models.AbstractModel):
    _name = "toomde.ohada.dsf.ml"
    _inherit = "toomde.ohada.annual.export.base"
    _description = "DSF Mali annual export (COMP-9 — DGI ML)"

    def _title(self):
        return "Déclaration Statistique et Fiscale — Mali (DGI)"

    def _sheet_name(self):
        return "DSF Mali"

    @api.model
    def _rows(self, company, year):
        rows = _ohada_common_rows(self, company, year)
        # Mali DGI includes a 7.5% withholding on services from
        # non-residents — declared annually alongside TVA.
        rows.append(("", ""))
        rows.append(("TVA collectée (compte 4431)", -self._prefix_total(company, year, "4431")))
        rows.append(("TVA déductible (compte 4452)", self._prefix_total(company, year, "4452")))
        rows.append(("Retenues sur prestations (compte 4424)", -self._prefix_total(company, year, "4424")))
        rows.append(("Impôt sur les sociétés (compte 891)", self._prefix_total(company, year, "891")))
        return rows


# ---------------------------------------------------------------------------
# Per-country DSF — Cameroun (DGI)
# ---------------------------------------------------------------------------

class DsfCamerounExport(models.AbstractModel):
    _name = "toomde.ohada.dsf.cm"
    _inherit = "toomde.ohada.annual.export.base"
    _description = "DSF Cameroun annual export (COMP-9 — DGI CM)"

    def _title(self):
        return "Déclaration Statistique et Fiscale — Cameroun (DGI)"

    def _sheet_name(self):
        return "DSF Cameroun"

    @api.model
    def _rows(self, company, year):
        rows = _ohada_common_rows(self, company, year)
        # CEMAC TVA is 19.25% (highest in OHADA) — single line + the
        # 5.5% withholding on services to non-residents.
        rows.append(("", ""))
        rows.append(("TVA collectée 19.25% (compte 4431)", -self._prefix_total(company, year, "4431")))
        rows.append(("TVA déductible (compte 4452)", self._prefix_total(company, year, "4452")))
        rows.append(("Retenues 5.5% sur prestations non-résidents (compte 4424)", -self._prefix_total(company, year, "4424")))
        rows.append(("Impôt sur les sociétés (compte 891)", self._prefix_total(company, year, "891")))
        return rows


# ---------------------------------------------------------------------------
# Per-country DSF — Niger (DGI)
# ---------------------------------------------------------------------------

class DsfNigerExport(models.AbstractModel):
    _name = "toomde.ohada.dsf.ne"
    _inherit = "toomde.ohada.annual.export.base"
    _description = "DSF Niger annual export (COMP-9 — DGI NE)"

    def _title(self):
        return "Déclaration Statistique et Fiscale — Niger (DGI)"

    def _sheet_name(self):
        return "DSF Niger"

    @api.model
    def _rows(self, company, year):
        rows = _ohada_common_rows(self, company, year)
        # Niger TVA is 19% (highest in UEMOA, alongside Guinea-Bissau).
        rows.append(("", ""))
        rows.append(("TVA collectée 19% (compte 4431)", -self._prefix_total(company, year, "4431")))
        rows.append(("TVA déductible (compte 4452)", self._prefix_total(company, year, "4452")))
        rows.append(("Impôt sur les bénéfices (compte 891)", self._prefix_total(company, year, "891")))
        return rows


# ---------------------------------------------------------------------------
# Per-country DSF — Togo (OTR)
# ---------------------------------------------------------------------------

class DsfTogoExport(models.AbstractModel):
    _name = "toomde.ohada.dsf.tg"
    _inherit = "toomde.ohada.annual.export.base"
    _description = "DSF Togo annual export (COMP-9 — OTR)"

    def _title(self):
        return "Déclaration Statistique et Fiscale — Togo (OTR)"

    def _sheet_name(self):
        return "DSF Togo"

    @api.model
    def _rows(self, company, year):
        rows = _ohada_common_rows(self, company, year)
        # Togo's authority is the OTR (Office Togolais des Recettes,
        # consolidating tax + customs since 2014). Standard 18% TVA;
        # no specific national withholding to break out.
        rows.append(("", ""))
        rows.append(("TVA collectée 18% (compte 4431)", -self._prefix_total(company, year, "4431")))
        rows.append(("TVA déductible (compte 4452)", self._prefix_total(company, year, "4452")))
        rows.append(("Impôt sur les sociétés (compte 891)", self._prefix_total(company, year, "891")))
        return rows


# ---------------------------------------------------------------------------
# Per-country DSF — Bénin (DGI)
# ---------------------------------------------------------------------------

class DsfBeninExport(models.AbstractModel):
    _name = "toomde.ohada.dsf.bj"
    _inherit = "toomde.ohada.annual.export.base"
    _description = "DSF Bénin annual export (COMP-9 — DGI BJ)"

    def _title(self):
        return "Déclaration Statistique et Fiscale — Bénin (DGI)"

    def _sheet_name(self):
        return "DSF Bénin"

    @api.model
    def _rows(self, company, year):
        rows = _ohada_common_rows(self, company, year)
        rows.append(("", ""))
        rows.append(("TVA collectée 18% (compte 4431)", -self._prefix_total(company, year, "4431")))
        rows.append(("TVA déductible (compte 4452)", self._prefix_total(company, year, "4452")))
        rows.append(("Impôt sur les sociétés (compte 891)", self._prefix_total(company, year, "891")))
        return rows


# ---------------------------------------------------------------------------
# Per-country DSF — Guinée-Bissau (DGCI)
# ---------------------------------------------------------------------------

class DsfGuineeBissauExport(models.AbstractModel):
    _name = "toomde.ohada.dsf.gw"
    _inherit = "toomde.ohada.annual.export.base"
    _description = "DSF Guinée-Bissau annual export (COMP-9 — DGCI GW)"

    def _title(self):
        return "Déclaration Statistique et Fiscale — Guinée-Bissau (DGCI)"

    def _sheet_name(self):
        return "DSF Guinée-Bissau"

    @api.model
    def _rows(self, company, year):
        rows = _ohada_common_rows(self, company, year)
        # Guinea-Bissau TVA is 19% (along with Niger, the highest UEMOA
        # rate). Authority: Direção-Geral das Contribuições e Impostos.
        rows.append(("", ""))
        rows.append(("TVA collectée 19% (compte 4431)", -self._prefix_total(company, year, "4431")))
        rows.append(("TVA déductible (compte 4452)", self._prefix_total(company, year, "4452")))
        rows.append(("Impôt sur les bénéfices (compte 891)", self._prefix_total(company, year, "891")))
        return rows


# ---------------------------------------------------------------------------
# Per-country DSF — Gabon (DGI)
# ---------------------------------------------------------------------------

class DsfGabonExport(models.AbstractModel):
    _name = "toomde.ohada.dsf.ga"
    _inherit = "toomde.ohada.annual.export.base"
    _description = "DSF Gabon annual export (COMP-9 — DGI GA)"

    def _title(self):
        return "Déclaration Statistique et Fiscale — Gabon (DGI)"

    def _sheet_name(self):
        return "DSF Gabon"

    @api.model
    def _rows(self, company, year):
        rows = _ohada_common_rows(self, company, year)
        # Gabon TVA: 18% standard + 10% reduced rate (basic foodstuffs).
        # Split TVA into two collected rows so DGI can cross-check.
        rows.append(("", ""))
        rows.append(("TVA collectée 18% (compte 44311)", -self._prefix_total(company, year, "44311")))
        rows.append(("TVA collectée 10% (compte 44312)", -self._prefix_total(company, year, "44312")))
        rows.append(("TVA déductible (compte 4452)", self._prefix_total(company, year, "4452")))
        rows.append(("Impôt sur les sociétés (compte 891)", self._prefix_total(company, year, "891")))
        return rows


# ---------------------------------------------------------------------------
# Per-country DSF — RDC (DGI)
# ---------------------------------------------------------------------------

class DsfRdcExport(models.AbstractModel):
    _name = "toomde.ohada.dsf.cd"
    _inherit = "toomde.ohada.annual.export.base"
    _description = "DSF RDC annual export (COMP-9 — DGI CD)"

    def _title(self):
        return "Déclaration Statistique et Fiscale — RD Congo (DGI)"

    def _sheet_name(self):
        return "DSF RDC"

    @api.model
    def _rows(self, company, year):
        rows = _ohada_common_rows(self, company, year)
        # DR Congo TVA is 16% (lowest in OHADA). Currency is CDF — not
        # XOF / XAF. Withholding on services paid to non-residents is
        # 14%, the highest in the zone — broken out separately.
        rows.append(("", ""))
        rows.append(("TVA collectée 16% (compte 4431)", -self._prefix_total(company, year, "4431")))
        rows.append(("TVA déductible (compte 4452)", self._prefix_total(company, year, "4452")))
        rows.append(("Retenues 14% sur services non-résidents (compte 4424)", -self._prefix_total(company, year, "4424")))
        rows.append(("Impôt sur les sociétés (compte 891)", self._prefix_total(company, year, "891")))
        return rows


# ---------------------------------------------------------------------------
# OHADA-wide Form 1004
# ---------------------------------------------------------------------------

class Form1004Export(models.AbstractModel):
    _name = "toomde.ohada.form1004"
    _inherit = "toomde.ohada.annual.export.base"
    _description = "OHADA Form 1004 annual export (COMP-10)"

    def _title(self):
        return "Formulaire OHADA 1004 — États financiers annuels"

    def _sheet_name(self):
        return "OHADA 1004"

    @api.model
    def _rows(self, company, year):
        totals = self._class_totals(company, year)
        # Form 1004 is OHADA-wide and pulls the same class totals into
        # the standardized layout. Same caveat as DSF — final cell map
        # to be aligned during integration testing with a real audit.
        return [
            ("Total Actif (Class 2 + 3 + 4D + 5D)",
             totals["2"] + totals["3"] + max(0, totals["4"]) + max(0, totals["5"])),
            ("Total Passif (Class 1 + 4C + 5C + 8)",
             -totals["1"] + min(0, totals["4"]) + min(0, totals["5"]) - totals["8"]),
            ("Total Charges (Class 6)", totals["6"]),
            ("Total Produits (Class 7)", -totals["7"]),
            ("Résultat de l'exercice (Class 8)", -totals["8"]),
        ]
